import openpyxl as xl
		
wb     = xl.load_workbook("file.xlsx")
sheet1 = wb["Sheet1"]



#.    1   STUDET DETAILS


		
def getStdntDtl(rawNum):
	
	lst=[]
	for i in range(1,5):
		cell=sheet1.cell(rawNum,i)
		celData = cell.value
		lst.append(celData)
#	print(lst[2])
	return lst

		
		
#_______________________________________________


#MARKS WITH ABSENT and TOTAL OK

		
def getMark(rawNum):
	
	lst=[]
	ttl=0
	for i in range(5,13):
		cell=sheet1.cell(rawNum,i)
		celData = cell.value
		if celData == None :
			lst.append("Absent")
		else :
			lst.append(int(celData))
			
	for item in lst:
			if item != "Absent" :
				ttl+=int(item)
	lst.append(ttl)
	return lst


#_______________________________________________


#_______________________________________________

#      MARKLIST WITHOUT ABSENT


def mrklst(rawNum):
	x = getMark(rawNum)
	n = 0
	for each in x:
		if each == "Absent":
			x[n]= 0
		n+=1

	return x



#_______________________________________________


#.   3.   PERCENTAGE

		
def getPercent(rawNum):
	marks = mrklst(rawNum)
#	print(marks,"mrks")
	lst=[]
	
	for i in range(8):
		lst.append(int(marks[i])*5)
	lst.append(int(int(marks[-1])/1.8))
	
	return lst
	